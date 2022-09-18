"""

    """

import json

import openpyxl as pyxl
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from update import fc


def create_wb_from_df(df) :
    wb = pyxl.Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df , index = False , header = True) :
        ws.append(r)
    return wb

def merge_cells_in_ws_col(ws , i_coln_tup) :
    mylist = []
    for cell in ws[i_coln_tup[1]] :
        mylist.append(cell.value)
    mergecount = 0
    startcell = 1
    color_i = 0
    for row in range(1 , len(mylist)) :
        if mylist[row - 1] == mylist[row] :
            mergecount += 1
        else :
            if mergecount > 0 :
                ws.merge_cells(start_row = startcell ,
                               start_column = i_coln_tup[0] ,
                               end_row = startcell + mergecount ,
                               end_column = i_coln_tup[0])

                ptf = PatternFill(start_color = colors[color_i] ,
                                  end_color = colors[color_i] ,
                                  fill_type = 'solid')
                ws.cell(row = startcell , column = i_coln_tup[0]).fill = ptf

                color_i += 1
                if color_i == len(colors) :
                    color_i = 0

            mergecount = 0
            startcell = row + 1
    if mergecount > 0 :
        ws.merge_cells(start_row = startcell ,
                       start_column = i_coln_tup[0] ,
                       end_row = startcell + mergecount ,
                       end_column = i_coln_tup[0])
        ptf = PatternFill(start_color = colors[color_i] ,
                          end_color = colors[color_i] ,
                          fill_type = 'solid')
        ws.cell(row = startcell , column = i_coln_tup[0]).fill = ptf
    return ws

def merge_cells_in_cols_in_ws(ws , coln_tups) :
    for coln_tup in coln_tups :
        ws = merge_cells_in_ws_col(ws , coln_tup)
    return ws

def make_all_cell_vertical_alignment_center(ws) :
    for row in ws.rows :
        for cell in row :
            cell.alignment = Alignment(vertical = 'center')
    return ws

def make_cells_in_cols_horizontal_alignment_center(ws , cols) :
    for col in cols :
        for cell in ws[col] :
            vert = cell.alignment.vertical
            cell.alignment = Alignment(horizontal = 'center' , vertical = vert)
    return ws

def make_all_cell_font_times_new_roman(ws) :
    for row in ws.rows :
        for cell in row :
            cell.font = Font(name = 'Times New Roman')
    return ws

def make_header_row_font_bigger(ws) :
    for cell in ws[1] :
        cell.font = Font(name = cell.font.name , bold = True , size = 20 , )
    return ws

def color_header_row(ws) :
    for cell in ws[1] :
        cell.fill = PatternFill(start_color = 'FCD5B4' ,
                                end_color = 'FCD5B4' ,
                                fill_type = 'solid')
    return ws

def make_header_row_bigger(ws) :
    ws.row_dimensions[1].height = 35
    return ws

def make_cols_font_bigger(ws , col_fsize_tups) :
    for col , fs in col_fsize_tups :
        for cell in ws[col][1 :] :
            cell.font = Font(name = cell.font.name , size = fs)
    return ws

colors = {
        0 : 'C0504D' ,
        1 : '538DD5' ,
        2 : 'C4D79B' ,
        3 : 'B1A0C7' ,
        4 : 'FABF8F' ,
        }

def color_rows(ws , start_col) :
    for col in list(ws.columns)[start_col :] :
        for i , cell in enumerate(col[1 :]) :
            if i % 2 == 0 :
                cell.fill = PatternFill(start_color = 'DCE6F1' ,
                                        end_color = 'DCE6F1' ,
                                        fill_type = 'solid')
    return ws

def fix_column_width(ws) :
    for column in ws.columns :
        length = max(len(str(cell.value)) for cell in column) * 1.23
        ws.column_dimensions[column[0].column_letter].width = length
    return ws

def fix_rows_width(ws) :
    for i in range(2 , len(list(ws.rows))) :
        ws.row_dimensions[i].height = 22
    return ws

def freeze_panes(ws , max_col_ltr) :
    ws.freeze_panes = f'{max_col_ltr}2'
    return ws

def main() :
    pass

    ##
    df = pd.read_excel('update.xlsx')
    ##
    with open('max.json' , 'r') as f :
        max_all_dc = json.load(f)
    max_all = int(max_all_dc['max'])
    ##
    cols_sort = [fc._1st_lvl , fc._2nd_lvl]
    cols_sort += list(range(3 , max_all))
    cols_sort += [fc.foln]
    ##
    df = df.sort_values(cols_sort)
    ##
    eng_ltr = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    cols_xl = [(i + 1 , x) for i , x in enumerate(eng_ltr[:max_all - 1])]
    ##
    wb = create_wb_from_df(df)
    ws = wb.active
    ##
    ws = merge_cells_in_cols_in_ws(ws , cols_xl)
    ##
    ws = make_all_cell_font_times_new_roman(ws)
    ##
    ws = make_all_cell_vertical_alignment_center(ws)
    ##
    cols_cent = [x for x in eng_ltr[:max_all - 1]]
    cols_cent += [x for x in eng_ltr[max_all : max_all + 2]]

    ws = make_cells_in_cols_horizontal_alignment_center(ws , cols_cent)
    ##
    ws = make_header_row_font_bigger(ws)
    ##
    ws = color_header_row(ws)
    ##
    ws = make_header_row_bigger(ws)
    ##
    cols_fsize = [('A' , 20) , ('B' , 18)]
    cols_fsize += [(x , 14) for x in eng_ltr[2 : len(df.columns)]]

    ws = make_cols_font_bigger(ws , cols_fsize)
    ##
    ws = color_rows(ws , max_all - 1)
    ##
    ws = fix_rows_width(ws)
    ##
    ws = fix_column_width(ws)
    ##
    ws = freeze_panes(ws , eng_ltr[max_all])
    ##
    wb.save('final.xlsx')

    ##

##
if __name__ == "__main__" :
    main()
    print('Done!')
